from openpyxl import load_workbook
from lxml import etree as ET
wb = load_workbook('cn8_2022_selfexptext.xlsx')
ws = wb.active
odoo = ET.Element("odoo")
root = ET.SubElement(odoo,"data")
row_num = 0
row_start = 2
for row in ws.iter_rows():
        if row_num != 2:
            row_num = row_num + 1 
            continue
        
        line_dict = {}
        key = 0
        for cell in row:
            line_dict[key] = cell.value
            key +=1
        record = ET.SubElement(root, "record")
        record.set('id',f"intrastat_category_2022_{line_dict[0]}" )
        record.set('model',"report.intrastat.code")

        notes = ET.SubElement(record, "field")
        notes.text = str(line_dict[0])
        notes.set('name',"name")

        description = ET.SubElement(record, "field")
        description.text = str(line_dict[1])
        description.set('name',"description")

        if line_dict[2] != None:
            quantity = ET.SubElement(record, "field")
            quantity.text = str(line_dict[2])
            quantity.set('name',"quantity")
    

tree = ET.ElementTree(root)
tree.write("intrastat_code.xml",encoding="UTF-8",pretty_print=True)

wb = load_workbook('cn8_2022_selfexptext.xlsx')
ws = wb.active
odoo = ET.Element("odoo")
root = ET.SubElement(odoo,"data")
row_num = 0
row_start = 2
for row in ws.iter_rows():
        if row_num != 2:
            row_num = row_num + 1 
            continue
        
        line_dict = {}
        key = 0
        for cell in row:
            line_dict[key] = cell.value
            key +=1
        record = ET.SubElement(root, "record")
        record.set('id',f"hs_code_2022_{line_dict[0]}" )
        record.set('model',"hs.code")

        notes = ET.SubElement(record, "field")
        notes.text = str(line_dict[0])
        notes.set('name',"local_code")

        description = ET.SubElement(record, "field")
        description.text = str(line_dict[1])
        description.set('name',"description")

        if line_dict[2] != None:
            quantity = ET.SubElement(record, "field")
            quantity.text = str(line_dict[2])
            quantity.set('name',"quantity")
    
        company_id = ET.SubElement(record, "field")
        company_id.set('name',"company_id")
        company_id.set('eval',"False")

tree = ET.ElementTree(root)
tree.write("intrastat_code_oca.xml",encoding="UTF-8",pretty_print=True)
