# -*- coding: utf-8 -*-
##############################################################################
#
#    Odoo, Open Source Enterprise Management Solution, third party addon
#    Copyright (C) 2018- Vertel AB (<http://vertel.se>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import except_orm, Warning, RedirectWarning
from odoo.addons.l10n_se_account_bank_statement_import.account_bank_statement_import import BankStatement


import logging
_logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from xlrd import open_workbook, XLRDError
from xlrd.book import Book
from xlrd.sheet import Sheet

import sys

class SEBKontohandelserrapport(object):
    """Parser for SEB Kontohändelser import files."""

    def __init__(self, data_file):
        """ Check if file can be read """
        try:
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (SEB Kontohändelser.xlsx)')
            raise ValueError(e)
        if (self.data.cell(5,1).value != 'Saldo' and self.data.cell(5,4).value != 'Reserverat belopp'):
            raise ValueError(u'This is not a SEB Kontohändelser document')

        self.nrows = self.data.nrows - 1
        self.header = []
        self.statements = []

    def parse(self):
        """Parse SEB transaktionsrapport bank statement file contents type 1."""

        _logger.info("Parsing SEB Kontohändlser")

        self.account_currency = 'SEK'
        self.header = [c.value.lower() for c in self.data.row(8)]
        self.account_number = self.data.cell(6,0).value
        self.name = self.data.cell(0,0).value[15:30]


        self.current_statement = BankStatement()
        self.current_statement.statement_id = f'STATEMENT:{self.account_number.replace(" ","")}'
        self.current_statement.date = fields.Date.today() # t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.start_balance = self.data.cell(self.nrows,5).value - self.data.cell(self.nrows,4).value
        self.current_statement.end_balance = self.data.cell(6, 1).value
        # self.current_statement.remote_owner = self.data.cell(15, 3).value

        for t in SEBIterator(self.data,header_row=8):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['belopp'])
            transaction.eref = t['verifikationsnummer'].strip()
            transaction.pref = t['text/mottagare'].strip()
            transaction.narration = t['text/mottagare'].strip()
            transaction.value_date = t['bokföringsdatum'] # bokföringsdatum,valutadatum
            transaction.unique_import_id = t['verifikationsnummer'].strip()
            transaction.remote_account = self.account_number.replace(" ", "");
            if transaction.name.startswith('Bg'):
                string = ' '.join(transaction.name.split())
                transaction.bg_account = string.split(' ')[1]
                transaction.bg_serial_number = string.split(' ')[2] if len(string.split(' ')) == 3 else ''

        self.statements.append(self.current_statement)
#        _logger.error('Statement %s Transaktioner %s' % (self.statements,''))  
        return self

        
class SEBFakturadetaljerDocumentParser(object):
    """Parser for SEB Fakturadetaljer import files."""

    def __init__(self, data_file):
        try:
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (SEB Kontohändelser.xlsx)')
            raise ValueError(e)
        if not (self.data.cell(0,0).value == u'Fakturadetaljer' and self.data.cell(3,0).value == u'Månad:'):
            _logger.error(u'Row 0 %s (was looking for Företagsnamn) %s %s' % (self.data.cell(0,0).value[:13],self.data.cell(3,0).value[:11],self.data.row(3)))
            raise ValueError(u'This is not a SEB Kontohändelser')

        self.nrows = self.data.nrows - 1
        self.header = []
        self.statements = []

    def parse(self):
        """Parse SEB Fakturadetaljer statement file contents"""

        _logger.info("Parsing SEB Kontohändlser from SEBtransaktionRaportType1")

        self.account_currency = 'SEK'
        self.header = [c.value.lower() for c in self.data.row(17)]
        self.account_number = self.data.cell(6,0).value
        self.name = self.data.cell(0,0).value[15:30]

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today() # t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = '%s %s' % (self.data.cell(0,0).value[14:],self.data.cell(2,0).value[6:])
        self.current_statement.start_balance = 0
        self.current_statement.end_balance = self.data.cell(self.nrows, 6).value.replace(u'\xa0', '')
        self.current_statement.remote_owner = self.data.cell(15, 3).value
        for t in SEBIterator(self.data,header_row=17):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['belopp'])
            transaction.eref = t['verifikationsnummer'].strip()
            transaction.pref = t['text/mottagare'].strip()
            transaction.narration = t['text/mottagare'].strip()
            transaction.value_date = t['bokfört'] # bokföringsdatum,valutadatum
            transaction.unique_import_id = t['verifikationsnummer'].strip()
            transaction.remote_owner = self.current_statement.remote_owner.strip() + t['datum']
            if transaction.name.startswith('Bg'):
                string = ' '.join(transaction.name.split())
                transaction.bg_account = string.split(' ')[1]
                transaction.bg_serial_number = string.split(' ')[2] if len(string.split(' ')) == 3 else ''

        self.statements.append(self.current_statement)
#        _logger.error('Statement %s Transaktioner %s' % (self.statements,''))  
        return self


class SEBTransaktionsrapportType2(object):
    """Parser for SEB Kontohändelser import files."""

    def __init__(self, data_file):
        try:
            #~ self.data_file = open_workbook(file_contents=data_file)
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (SEB Kontohändelser.xlsx)')
            raise ValueError(e)
        self.nrows = self.data.nrows - 1
        self.header = []
        self.statements = []
        if not (self.data.cell(2,0).value[:7] == u'Datum: ' and self.data.cell(4,0).value[:15] == u'Bokföringsdatum' and self.data.cell(4,3).value[:16] == u'text / mottagare'):
            _logger.error(u'Row 0 %s (was looking for Datum / Bokföringsdatum) %s %s' % (self.data.cell(2,0).value[:7],self.data.cell(4,0).value[:15],self.data.row(0)))
            raise ValueError(u'This is not a SEB Kontohändelser Typ2')


    def parse(self):
        """Parse SEB transaktionsrapport bank statement file contents type 2."""

        _logger.info("Parsing SEB Kontohändlser from SEBTRansaktionRaportType2")

        self.account_currency = 'SEK'
        self.header = []
        self.account_number = self.data.cell(1,0).value.strip()
        self.name = self.data.cell(0,0).value[15:30]

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today() # t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = '%s %s' % (self.data.cell(0,0).value,self.data.cell(2,0).value[6:])
        self.current_statement.start_balance = float(self.data.cell(self.nrows,5).value - self.data.cell(self.nrows,4).value)
        self.current_statement.end_balance = float(self.data.cell(5,5).value)
        for t in SEBIterator(self.data,header_row=4):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['belopp'])
            transaction.eref = t['verifikationsnummer'].strip()
            transaction.pref = t['text / mottagare'].strip()
            transaction.narration = t['text / mottagare'].strip()
            transaction.value_date = t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
            transaction.unique_import_id = t['verifikationsnummer'].strip()
            transaction.remote_owner = t['text / mottagare'].strip()

        self.statements.append(self.current_statement)
#        _logger.error('Statement %s Transaktioner %s' % (self.statements,''))
        return self


class SEBTransaktionsrapportType3(object):
    """Parser for SEB Kontohändelser import files."""

    def __init__(self, data_file):
        try:
            #~ self.data_file = open_workbook(file_contents=data_file)
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (SEB Kontohändelser.xlsx)')
            raise ValueError(e)
        self.nrows = self.data.nrows - 1
        self.header = []
        self.statements = []
        if not (self.data.cell(2,0).value[:7] == u'Datum: ' and self.data.cell(4,0).value[:15] == u'Bokföringsdatum' and self.data.cell(4,3).value[:14] == u'Text/mottagare'):
            _logger.error(u'Row 0 %s (was looking for Datum / Bokföringsdatum) %s %s' % (self.data.cell(2,0).value[:7],self.data.cell(4,0).value[:15],self.data.row(0)))
            raise ValueError(u'This is not a SEB Kontohändelser Typ3')


    def parse(self):
        """Parse SEB transaktionsrapport bank statement file contents type 3."""

        _logger.info("Parsing SEB Kontohändlser from SEBTRansaktionRaportType3")

        self.account_currency = 'SEK'
        self.header = []
        self.account_number = self.data.cell(1,0).value.strip()
        self.name = self.data.cell(0,0).value[15:30]

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today() # t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = '%s %s' % (self.data.cell(0,0).value,self.data.cell(2,0).value[6:])
        self.current_statement.start_balance = float(self.data.cell(self.nrows,5).value - self.data.cell(self.nrows,4).value)
        self.current_statement.end_balance = float(self.data.cell(5,5).value)
        for t in SEBIterator(self.data,header_row=4):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['belopp'])
            transaction.eref = t['verifikationsnummer'].strip()
            transaction.pref = t['text/mottagare'].strip()
            transaction.partner_name = t['text/mottagare'].strip()
            transaction.narration = t['text/mottagare'].strip()
            transaction.value_date = t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
            transaction.unique_import_id = t['verifikationsnummer'].strip()
            transaction.remote_owner = t['text/mottagare'].strip()

            #~ transaction.message
            #self.current_statement.end_balance =

        self.statements.append(self.current_statement)
#        _logger.error('Statement %s Transaktioner %s' % (self.statements,''))
        return self



class SEBTransaktionsrapportType4(object):
    """Parser for SEB Kontohändelser import files."""

    def __init__(self, data_file):
        try:
            #~ self.data_file = open_workbook(file_contents=data_file)
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (SEB Kontohändelser.xlsx)')
            raise ValueError(e)
        self.nrows = self.data.nrows - 1
        self.header = []
        self.statements = []
        if not (self.data.cell(1,1).value[:13] == u'Transaktioner' and self.data.cell(4,1).value[:13] == u'Företagskonto' and self.data.cell(2,1).value[:14] == u'Business Arena'):
            _logger.error(u'Row 0 %s (was looking for Transaktioner / Business Arena) %s %s' % (self.data.cell(1,1).value[:13],self.data.cell(2,1).value[:14],self.data.row(0)))
            raise ValueError(u'This is not a SEB Kontohändelser Typ4')


    def parse(self):
        """Parse SEB transaktionsrapport bank statement file contents type 4."""

        _logger.info("Parsing SEB Kontohändlser from SEBTRansaktionRaportType4")

        self.account_currency = 'SEK'
        self.header = []
        self.account_number = self.data.cell(4,1).value[14:14].strip()
        self.name = self.data.cell(5,1).value[13:30].strip()

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today() # t[u'bokföringsdatum'] # bokföringsdatum,valutadatum
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = '%s %s' % (self.data.cell(5,1).value[13:],self.data.cell(6,1).value[8:])
        #self.current_statement.start_balance = float(self.data.cell(self.nrows,5).value - self.data.cell(self.nrows,4).value)
        self.current_statement.end_balance = float(self.data.cell(9,7).value)
        for t in SEBIterator(self.data,header_row=8):
            transaction = self.current_statement.create_transaction()
            if isinstance(t['insättningar'], str):
                transaction.transferred_amount = float(t['uttag'])
            else:
                transaction.transferred_amount = float(t['insättningar'])
            transaction.eref = t['text'].strip() + t['typ'].strip() 
            transaction.pref = t['typ'].strip()
            transaction.partner_name = t['text'].strip()
            transaction.narration = t['text'].strip()
            transaction.value_date = t[u'bokförd'] 
            #transaction.unique_import_id = t['verifikationsnummer'].strip()
            transaction.remote_owner = t['text'].strip()

            transaction.message = t['text'].strip()
            #self.current_statement.end_balance =

        self.statements.append(self.current_statement)
#        _logger.error('Statement %s Transaktioner %s' % (self.statements,''))
        return self



class SEBIterator(object):
    def __init__(self, data,header_row=8):
        self.row = header_row + 1 # First data-row
        self.data = data
        self.header = [c.value.lower() for c in data.row(header_row)]

    def __iter__(self):
        return self

    def next(self):
        if self.row >= self.data.nrows:
            raise StopIteration
        r = self.data.row(self.row)
        self.row += 1
        return {self.header[n]: r[n].value for n in range(len(self.header))}

    def __next__(self):
        if self.row >= self.data.nrows:
            raise StopIteration
        r = self.data.row(self.row)
        self.row += 1
        return {self.header[n]: r[n].value for n in range(len(self.header))}

# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
