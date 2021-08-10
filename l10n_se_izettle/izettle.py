# -*- coding: utf-8 -*-
##############################################################################
#
#    Odoo, Open Source Enterprise Management Solution, third party addon
#    Copyright (C) 2021- Vertel AB (<https://vertel.se>).
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU Affero General Public License as
#    published by the Free Software Foundation, either version 3 of the
#    License, or (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU Affero General Public License for more details.
#
#    You should have received a copy of the GNU Affero General Public License
#    along with this program.  If not, see <http://www.gnu.org/licenses/>.
#
##############################################################################
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import except_orm, Warning, RedirectWarning
from odoo.addons.l10n_se_account_bank_statement_import.account_bank_statement_import import BankStatement

import logging
_logger = logging.getLogger(__name__)

from openpyxl import load_workbook
from xlrd import open_workbook, XLRDError
from xlrd.book import Book
from xlrd.sheet import Sheet
from datetime import datetime

import sys


class IzettleTransaktionsrapportXlsType(object):
    """Parser for iZettle Kontohändelser import files."""

    def __init__(self, data_file):
        try:
            #~ self.data_file = open_workbook(file_contents=data_file)
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (iZettle Kontohändelser.xls)')
            raise ValueError(e)
        if not (self.data.cell(5,0).value[:20] == u'Betalningsmottagare:' and self.data.cell(10,0).value[:21] == u'Betalningsförmedlare:' and self.data.cell(16,8).value == u'Korttyp'):
            _logger.error(u'Header did not contain "Betalningsmottagare:" and "Betalningsförmedlare:" and "Korttyp".')
            raise ValueError(u'This is not a iZettle Report')

        self.nrows = self.data.nrows - 17
        self.header = []
        self.statements = []

    def parse(self):
        """Parse iZettle transaktionsrapport bank statement file type 1."""

        self.account_currency = 'SEK'
        self.header = [c.value.lower() for c in self.data.row(16)]
        self.account_number = self.data.cell(13,2).value
        self.name = self.data.cell(5,2).value

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today()
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = 'iZettle %s' % self.data.cell(3,2).value
        self.current_statement.start_balance = 0.0
        for t in IzettleIterator(self.data, header_row=16):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['netto'])
            transaction.original_amount = float(t['totalt'])
            self.current_statement.end_balance += float(t['netto'])
            transaction.eref = int(t['kvittonummer'])
            transaction.name = '%s %s' % (t['korttyp'].strip(),t['sista siffror'].strip())
            transaction.note = 'Totalt: %s\nMoms: %s\nAvgift: %s\n%s %s' % (float(t['totalt']), t['moms (25.0%)'], t['avgift'], t['korttyp'].strip(), t['sista siffror'].strip())
            transaction.value_date = t[u'datum']
            transaction.unique_import_id = int(t['kvittonummer'])

        self.statements.append(self.current_statement)
        return self

class IzettleXlrdTransaktionsrapportXlsxType(object):
    """Parser for iZettle Kontohändelser import files."""

    def __init__(self, data_file):
        try:
            #~ self.data_file = open_workbook(file_contents=data_file)
            self.data = open_workbook(file_contents=data_file).sheet_by_index(0)
        except XLRDError as e:
            _logger.error(u'Could not read file (iZettle Kontohändelser.xls)')
            raise ValueError(e)
        if not (self.data.cell(5,0).value[:20] == u'Betalningsmottagare:' and self.data.cell(10,0).value[:21] == u'Betalningsförmedlare:' and self.data.cell(16,8).value == u'Kortutgivare'):
            _logger.error(u'Header did not contain "Betalningsmottagare:" and "Betalningsförmedlare:" and "Korttyp".')
            raise ValueError(u'This is not a iZettle Report')

        self.nrows = self.data.nrows - 17
        self.header = []
        self.statements = []

    def floatHourToTime(self,fh):
        hours, hourSeconds = divmod(fh, 1)
        minutes, seconds = divmod(hourSeconds * 60, 1)
        return (
            int(hours),
            int(minutes),
            int(seconds * 60),
        )
        
    def parse(self):
        """Parse iZettle transaktionsrapport bank statement file type 1."""

        self.account_currency = 'SEK'
        self.header = [c.value.lower() for c in self.data.row(16)]
        self.account_number = self.data.cell(13,2).value
        self.name = self.data.cell(5,2).value

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today()
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = 'iZettle %s' % self.data.cell(3,2).value
        self.current_statement.start_balance = 0.0
        for t in IzettleIterator(self.data, header_row=16):
            transaction = self.current_statement.create_transaction()
            transaction.transferred_amount = float(t['netto'])
            transaction.original_amount = float(t['totalt'])
            self.current_statement.end_balance += float(t['netto'])
            transaction.eref = int(t['kvittonummer'])
            transaction.name = '%s %s' % (t['kortutgivare'].strip(),t['sista siffror'].strip())
            transaction.note = 'Totalt: %s\nMoms: %s\nAvgift: %s\n%s %s' % (float(t['totalt']), t['moms (25.0%)'], t['avgift'], t['kortutgivare'].strip(), t['sista siffror'].strip())
            
            dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(t[u'datum']) - 2)
            hour, minute, second = self.floatHourToTime(t[u'datum'] % 1)
            dt = dt.replace(hour=hour, minute=minute, second=second)
            _logger.warn("dt is {}".format(dt))
            
            transaction.value_date = dt   
            transaction.unique_import_id = int(t['kvittonummer'])

        self.statements.append(self.current_statement)
        return self
    
    

class IzettleTranskationReportXlsxType(object):
    """ Parser for iZettle Kontohändelser import files. """
    def __init__(self, data_file):
        try:
            wb = load_workbook(filename=BytesIO(data_file))
            ws = wb.get_sheet_names()
            _logger.debug(ws) #REQURIED? DELETE AT YOUR OWN RISK. 
            self.data = wb.get_sheet_by_name(ws[0])
        # TODO?: Catch BadZipFile, IOerror, ValueError
        except:
            raise ValueError('This is not a iZettle xlsx document')
        if not (self.data.cell(6,1).value[:20] == u'Betalningsmottagare:' and self.data.cell(11,1).value[:21] == u'Betalningsförmedlare:' and self.data.cell(17,9).value == u'Kortutgivare'):
            _logger.error(u'Header did not contain "Betalningsmottagare" or "Betalningsförmedlare" or "Kortutgivare".')
            raise ValueError(u'This is not a iZettle xlsx Report')
        
        self.header = []
        self.statements = []
        
    def parse(self):
        """ Parse iZettle transactionsreport bank statement file type 2. """
        
        self.account_currency = 'SEK'
        self.account_number = self.accountnumbergenerator() #TODO: Change to actual sheet value? There is none currently?
        self.name = self.data.cell(4,3).value

        self.current_statement = BankStatement()
        self.current_statement.date = fields.Date.today()
        self.current_statement.local_currency = self.account_currency or 'SEK'
        self.current_statement.local_account =  self.account_number
        self.current_statement.statement_id = 'iZettle %s' % self.data.cell(4,3).value
        self.current_statement.start_balance = 0.0
        self.current_statement.end_balance = 0.0
        
        for index, row in enumerate(self.data.iter_rows(17, self.data.max_row-3, values_only=True), start=17):
            if (index == 17):
                self.header = {c:i for i, c in enumerate(row)}
                _logger.warn(self.header)
            else:
                tdict = {key:row[self.header[key]] for key in self.header}
                transaction = self.current_statement.create_transaction()
                transaction['amount'] = tdict[u'Totalt']
                transaction['account_number'] = self.account_number
                transaction['original_amount'] = tdict[u'Totalt']
                _logger.warn("Tid is: {}".format(tdict['Tid']))
                transaction['date'] = tdict['Tid'].strftime("%Y-%m-%d")
                transaction['name'] = int(tdict[u'Kvittonummer'])
                transaction['note'] = 'Totalt: {}\nMoms: {}\nAvgift: {}\n{} {}'.format(tdict[u'Totalt'], tdict[u'Moms (25.0%)'], tdict[u'Avgift'], tdict[u'Kortutgivare'].strip(), tdict[u'Sista siffror'].strip())
                self.current_statement.end_balance += tdict[u'Totalt']
                
        self.statements.append(self.current_statement)
        return self
    
    def accountnumbergenerator(self):
        def accn(part):
            return ''.join([ i for i in part if i.isdigit()])
        result = '-'.join([accn(self.data.cell(12,3).value), accn(self.data.cell(13,3).value), accn(self.data.cell(14,3).value)])
        return result
    

class IzettleIterator(object):
    def __init__(self, data,header_row=16):
        self.row = header_row + 1
        self.data = data
        self.header = [c.value.lower() for c in data.row(header_row)]

    def __iter__(self):
        return self

    def next(self):
        if self.row >= self.data.nrows - 3:
            raise StopIteration
        r = self.data.row(self.row)
        self.row += 1
        return {self.header[n]: r[n].value for n in range(len(self.header))}

# vim:expandtab:smartindent:tabstop=4:softtabstop=4:shiftwidth=4:
