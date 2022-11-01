from openpyxl import load_workbook
from lxml import etree as ET
wb = load_workbook('INK2_19-P1-exkl-version-2022-11-1.xlsx')
ws = wb.active
odoo = ET.Element("odoo")
root = ET.SubElement(odoo,"data")

previous_rad_ink_2 = ""
previous_benamning = ""

for row in ws.iter_rows():
   cellstr = ""
   # ~ print(row[3])
   line_dict = {}
   key = 0
   for cell in row:
     line_dict[key] = cell.value
     key +=1
   if type(line_dict[0]) is int and len(str(line_dict[0])) == 4:
      # ~ print(line_dict[3])
      # ~ subi.set(attr, value)
      record = ET.SubElement(root, "record")
      # ~ id="account_sru_code_7527" model="account.sru.code"
      record.set('id',f"account_sru_code_{line_dict[0]}" )
      record.set('model',"account.sru.code")
      sru_code = ET.SubElement(record, "field")
      sru_code.text = str(line_dict[0])
      sru_code.set('name',"sru_code")
      
      notes = ET.SubElement(record, "field")
      notes.text = str(line_dict[6])
      notes.set('name',"notes")
      
      
      rad_ink_2 = ET.SubElement(record, "field")
      if line_dict[1] == None:
        rad_ink_2.text = previous_rad_ink_2
      else:
        rad_ink_2.text = str(line_dict[1])
        previous_rad_ink_2 = str(line_dict[1])
      rad_ink_2.set('name',"rad_ink_2")
      
      benamning = ET.SubElement(record, "field")
      if line_dict[2] == None:
        benamning.text = previous_benamning
      else:
        benamning.text = str(line_dict[2])
        previous_benamning = str(line_dict[2])
      benamning.set('name',"benamning")
      
      if line_dict[3] != None:
          text_intervall_original = ET.SubElement(record, "field")
          text_intervall_original.text = str(line_dict[3])
          text_intervall_original.set('name',"text_intervall_original")
          if line_dict[4] != None:
            text_intervall = ET.SubElement(record, "field")
            text_intervall.text = str(line_dict[4]).replace(".",",")
            text_intervall.set('name',"text_intervall")
          if line_dict[5] != None:
            text_intervall_exclude = ET.SubElement(record, "field")
            text_intervall_exclude.text = str(line_dict[5]).replace(".",",")
            text_intervall_exclude.set('name',"text_intervall_exclude")

tree = ET.ElementTree(root)
tree.write("sru_account.xml",encoding="UTF-8",pretty_print=True)
    
